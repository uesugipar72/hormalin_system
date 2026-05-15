import tkinter as tk
from tkinter import ttk
from services.history_service import get_history
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import os
import win32com.client
from openpyxl.styles import Font, Border, Side, Alignment
import time
import pythoncom
import tempfile
import uuid
from tkcalendar import Calendar
from datetime import datetime, timedelta
from gui.base_frame import BaseFrame

class HistoryFrame(BaseFrame):

    window_size = "1200x800"
    resizable = (True,True)
   
    def __init__(self, parent, controller):

        super().__init__(parent,controller)

        ttk.Label(self, text="取引履歴画面").pack(pady=10)

        # ▼ 抽出コンボボックス
        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill="x", padx=10, pady=5, anchor="w")

        ttk.Label(filter_frame, text="表示抽出").pack(side="left", padx=5)

        self.filter_cb = ttk.Combobox(
            filter_frame,
            values=[
                "すべて",
                "8mLホルマリン",
                "100mLホルマリン",
                "150mLホルマリン"
            ],
            state="readonly",
            width=20
        )
        self.filter_cb.set("すべて")
        self.filter_cb.pack(side="left", padx=5)


        ttk.Label(filter_frame, text="開始日").pack(side="left", padx=(15, 5))
        today = datetime.now()
        three_months_ago = today - timedelta(days=90)

        ttk.Label(filter_frame, text="開始日").pack(side="left", padx=(15, 5))

        # =====================================
        # 開始日
        # =====================================

        ttk.Label(filter_frame, text="開始日").pack(side="left", padx=(15, 5))

        self.start_date_var = tk.StringVar()
        self.start_date_var.set(three_months_ago.strftime("%Y-%m-%d"))

        self.start_date_entry = ttk.Entry(
            filter_frame,
            textvariable=self.start_date_var,
            width=12
        )
        self.start_date_entry.pack(side="left")

        ttk.Button(
            filter_frame,
            text="📅",
            width=3,
            command=lambda: self.open_calendar(self.start_date_var)
        ).pack(side="left", padx=(2, 10))


        # =====================================
        # 終了日
        # =====================================

        ttk.Label(filter_frame, text="終了日").pack(side="left", padx=(15, 5))

        self.end_date_var = tk.StringVar()
        self.end_date_var.set(today.strftime("%Y-%m-%d"))

        self.end_date_entry = ttk.Entry(
            filter_frame,
            textvariable=self.end_date_var,
            width=12
        )
        self.end_date_entry.pack(side="left")

        ttk.Button(
            filter_frame,
            text="📅",
            width=3,
            command=lambda: self.open_calendar(self.end_date_var)
        ).pack(side="left", padx=(2, 10))

       

        # ▼ 抽出ボタン
        tk.Button(
            filter_frame,
            text="抽出",
            command=self.refresh,
            bg="#0078D7",
            fg="white",
            activebackground="#1E90FF",
            activeforeground="white",
            font=("Meiryo", 10, "bold")
        ).pack(side="left", padx=5)

        # ▼ 戻るボタン
        ttk.Button(
            filter_frame,
            text="メニューに戻る",
            command=lambda: controller.show_frame("MenuFrame")
        ).pack(side="left",padx=10)

        # ▼ Excel出力ボタン
        ttk.Button(
            filter_frame,
            text="Excel出力",
            command=self.export_excel
        ).pack(side="left",padx=5)

        ttk.Button(
            filter_frame,
            text="PDF出力",
            command=self.export_pdf
        ).pack(side="left",padx=5)

        

        # ▼ Treeview
        columns = ("日時", "ソート日時","ホルマリン種", "区分", "数量", "担当者", "在庫", "備考")

        self.tree = ttk.Treeview(self, columns=columns, show="headings")

        # ▼ 行カラー設定（追加）
        self.tree.tag_configure("evenrow", background="#F5F5F5")
        self.tree.tag_configure("oddrow", background="white")

        self.tree.column("日時", width=160,anchor="center")
        self.tree.column("ソート日時", width=0,stretch=False)
        self.tree.column("ホルマリン種", width=160,anchor="center")
        self.tree.column("区分", width=60,anchor="center")
        self.tree.column("数量", width=50,anchor="e")
        self.tree.column("担当者", width=120,anchor="center")
        self.tree.column("在庫", width=50,anchor="e")
        self.tree.column("備考", width=200,anchor="w")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        self.sort_reverse = False  # 昇順/降順フラグ

        for col in columns:
            self.tree.heading(
                col,
                text=col,
                command=lambda c=col: self.sort_by_column(c)
            )

    def open_calendar(self, target_var):

        top = tk.Toplevel(self)

        top.title("日付選択")
        top.geometry("300x250")

        top.transient(self)
        top.grab_set()
        top.resizable(False, False)

        cal = Calendar(
            top,
            selectmode="day",
            date_pattern="yyyy-mm-dd"
        )

        cal.pack(fill="both", expand=True, padx=10, pady=10)

        # ▼ 現在値があれば反映
        try:
            current_date = datetime.strptime(
                target_var.get(),
                "%Y-%m-%d"
            )

            cal.selection_set(current_date)

        except:
            pass

        # ▼ 決定処理
        def select_date():

            target_var.set(cal.get_date())

            top.destroy()

        ttk.Button(
            top,
            text="決定",
            command=select_date
        ).pack(pady=5)   

    def sort_by_column(self, col):

        col_index = self.tree["columns"].index(col)

        data = []
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            data.append((values[col_index], item, values))

        # ▼ ここを中に入れる！！
        def convert(value, col, values):
              if col == "日時":

                    return datetime.strptime(
                        values[1],   # hidden ソート日時
                        "%Y-%m-%d %H:%M:%S"
                    )

              try:
                    return float(value)

              except:
                    return value

        data.sort(
            key=lambda x: convert(x[0], col, x[2]),
            reverse=self.sort_reverse
        )

        for index, (_, item, _) in enumerate(data):
            self.tree.move(item, "", index)

        self.sort_reverse = not self.sort_reverse

    def refresh(self):
        data = self.get_history_data()
        action_map = {
            "IN": "入庫",
            "OUT": "出庫"
        }

        selected = self.filter_cb.get()

        start_date_str = self.start_date_var.get().strip()
        end_date_str = self.end_date_var.get().strip()

        start_date = None
        end_date = None

        try:
            if start_date_str:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")

            if end_date_str:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

        except ValueError:
            messagebox.showerror(
                "入力エラー",
                "日付は YYYY-MM-DD 形式で入力してください"
            )
            return

        # ▼ 既存削除
        for row in self.tree.get_children():
            self.tree.delete(row)

        # ▼ 再表示（フィルタ付き）
        for row in sorted(
            data,
            key=lambda x: datetime.strptime(x[0], "%Y-%m-%d %H:%M:%S")
        ):
            row = list(row)

            # 区分変換
            dt = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
            original_datetime = row[0]
            date_str = row[0][:10]
            y, m, d = date_str.split("-")
            row[0] = dt.strftime("%Y年%m月%d日")

            row[2] = action_map.get(row[2], row[2])# 数量の小数点を消す
            row[3] = int(row[3]) if row[3] is not None else 0
            row[5] = int(row[5]) if row[5] is not None else 0
            #row[7] = int(row[7]) if row[7] is not None else 0

            # ▼ フィルタ処理
            if selected != "すべて":
                if row[1] != selected:
                    continue
            # ▼ 日付フィルタ
            row_date = datetime.strptime(date_str, "%Y-%m-%d")

            if start_date and row_date < start_date:
                continue

            if end_date and row_date > end_date:
                continue

            tag = "evenrow" if len(self.tree.get_children()) % 2 == 0 else "oddrow"

            self.tree.insert(
                "",
                "end",
                iid=str(uuid.uuid4()),
                values=(
                    row[0],              # 表示用
                    original_datetime,   # ソート用
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    row[6]
                ),
                tags=(tag,)
            )

    def get_history_data(self):
        return get_history()

    def export_excel(self):

        documents = os.path.join(
        os.path.expanduser("~"),
        "Documents"
        )

        excel_path = filedialog.asksaveasfilename(
            initialdir=documents,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Excel    保存先を選択"
        )
        if not excel_path:
            return

        wb = self.create_workbook()

        wb.save(excel_path)

        messagebox.showinfo(
            "完了",
            "Excelファイルを保存しました"
        )

        os.startfile(excel_path)

    # ▼ Excel → PDF変換
    def export_pdf(self):
        documents = os.path.join(
        os.path.expanduser("~"),
        "Documents"
        )

        pdf_path = filedialog.asksaveasfilename(
            initialdir=documents,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="PDF保存先を選択"
        )

        if not pdf_path:
            return

        excel_path = pdf_path.replace(".pdf", ".xlsx")

        excel = None
        wb_excel = None

        try:

            wb = self.create_workbook()
            wb.save(excel_path)

            pythoncom.CoInitialize()

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb_excel = excel.Workbooks.Open(
                os.path.abspath(excel_path)
            )

            ws_excel = wb_excel.Worksheets(1)

            ws_excel.ExportAsFixedFormat(
                0,
                os.path.abspath(pdf_path)
            )

            time.sleep(1)

            if os.path.exists(pdf_path):

                messagebox.showinfo(
                    "完了",
                    "PDFを出力しました"
                )

                os.startfile(pdf_path)

            else:

                messagebox.showerror(
                    "エラー",
                    "PDFファイルが見つかりません"
                )

        except Exception as e:

            messagebox.showerror(
                "PDF出力エラー",
                str(e)
            )

        finally:

            if wb_excel:
                wb_excel.Close(False)

            if excel:
                excel.Quit()

            if os.path.exists(excel_path):
                try:
                    os.remove(excel_path)
                except:
                    pass
    def create_workbook(self):

        wb = Workbook()
        ws = wb.active
        ws.title = "取引履歴"

        # =====================================
        # タイトル
        # =====================================

        ws.merge_cells("A1:G1")

        title_cell = ws["A1"]
        title_cell.value = "ホルマリン管理台帳"

        title_cell.font = Font(
            bold=True,
            size=16
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # =====================================
        # ヘッダー
        # =====================================

        headers = (
            "日時",
            "ホルマリン種",
            "区分",
            "数量",
            "担当者",
            "在庫",
            "備考"
        )

        for col_num, header in enumerate(headers, start=1):

            cell = ws.cell(row=3, column=col_num)

            cell.value = header

            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # =====================================
        # データ
        # =====================================

        start_row = 4

        for row_index, item in enumerate(
            self.tree.get_children(),
            start=start_row
        ):

            values = self.tree.item(item, "values")

            for col_index, value in enumerate(values, start=1):

                cell = ws.cell(
                    row=row_index,
                    column=col_index
                )

                cell.value = value

                # ▼ 列ごとの配置
                if col_index in [1, 2, 3, 5]:
                    # 中央寄せ
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                elif col_index in [4, 6]:
                    # 右寄せ
                    cell.alignment = Alignment(
                        horizontal="right",
                        vertical="center"
                    )

                else:
                    # 左寄せ
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center"
                    )

        # =====================================
        # 列幅
        # =====================================

        widths = {
            "A": 18,
            "B": 20,
            "C": 10,
            "D": 10,
            "E": 15,
            "F": 10,
            "G": 40
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # =====================================
        # 印刷設定
        # =====================================

        ws.print_title_rows = '1:3'

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # =====================================
        # 罫線
        # =====================================

        thin = Side(
            style="thin",
            color="000000"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(
            min_row=3,
            max_row=max_row,
            min_col=1,
            max_col=max_col
        ):

            for cell in row:
                cell.border = border

        return wb